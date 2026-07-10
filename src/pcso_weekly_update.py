#!/usr/bin/env python3
"""Validate and reproduce the registered PCSO weekly monitoring run.

This script does not fetch draws or edit the workbook. It validates the dated
provenance manifest against the append-only CSVs, protects the frozen
exploration prefixes, runs the standing m=9 monitoring family, and verifies
the workbook's structural invariants.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import random
from collections import Counter
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
DATASET_DIR = ROOT / "datasets" / "pcso-lotto"
DEFAULT_MANIFEST = (
    DATASET_DIR / "provenance" / "pcso_weekly_2026-07-08.json"
)
DEFAULT_OUTPUT = ROOT / "results" / "pcso_confirmation_2026-07-08.json"
WORKBOOK = ROOT / "PCSO_Lotto_Analysis_Mar-Jun_2026.xlsx"

GAMES = {
    "Lotto 6/42": 42,
    "Mega Lotto 6/45": 45,
    "Super Lotto 6/49": 49,
    "Grand Lotto 6/55": 55,
    "Ultra Lotto 6/58": 58,
}

DRAW_FILES = {
    "data_draws.csv": "Draw Date",
    "data_draws_1yr.csv": "Date",
    "data_draws_1yr_audited.csv": "Date",
}

EXPLORATION_PREFIX_SHA256 = {
    "data_draws.csv": (
        "8f38c1d2fc1f3778552f015b4dc690c603be4a0d8c3659d692691d2ceb5f5087"
    ),
    "data_draws_1yr.csv": (
        "0a9d06d08557bc6010d2a04415a3fb94ec696780f8214538afd4cd3c815c36b0"
    ),
    "data_draws_1yr_audited.csv": (
        "ddcaf22dc665d5837afe840bdb1816f11fdda4908ef0ff270ec091a477ce7693"
    ),
    "data_astro_geomagnetic.csv": (
        "ee511ace0a90f39afa105f40ab2e8f6e16bb6bfd3ccb626a9782fdde65ab9427"
    ),
}

ASTRO_FILE = "data_astro_geomagnetic.csv"
ASTRO_DATE = "Draw Date"
MOON_ALT = "Moon Alt (deg)"
MOON_ILLUM = "Moon Illum (0-1)"
MEAN_DRAWN = "Mean drawn # / pool"
KP_DRAW = "Kp at draw (geomagnetic)"
KP_DAILY = "Kp daily mean"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def is_date(value: str) -> bool:
    return (
        len(value) == 10
        and value[4] == "-"
        and value[7] == "-"
        and value[:4].isdigit()
    )


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def dated_rows(path: Path, date_column: str) -> list[dict[str, str]]:
    return [r for r in load_csv(path) if is_date(r.get(date_column, ""))]


def exploration_prefix_sha256(path: Path, date_column: str, cutoff: str) -> str:
    lines = path.read_bytes().splitlines(keepends=True)
    if not lines:
        raise ValueError(f"{path}: empty CSV")
    header = next(
        csv.reader([lines[0].decode("utf-8-sig").rstrip("\r\n")])
    )
    date_index = header.index(date_column)
    kept = [lines[0]]
    for raw_line in lines[1:]:
        row = next(csv.reader([raw_line.decode("utf-8").rstrip("\r\n")]))
        if len(row) <= date_index:
            continue
        date = row[date_index]
        if is_date(date) and date <= cutoff:
            kept.append(raw_line)
    return hashlib.sha256(b"".join(kept)).hexdigest()


def require_crlf(path: Path) -> None:
    raw = path.read_bytes()
    if raw.count(b"\n") != raw.count(b"\r\n"):
        raise ValueError(f"{path}: expected every line to use CRLF")


def draw_record(row: dict[str, str], date_column: str) -> dict[str, object]:
    return {
        "date": row[date_column],
        "game": row["Game"],
        "numbers": [int(row[f"N{i}"]) for i in range(1, 7)],
    }


def draw_key(record: dict[str, object]) -> tuple[str, str]:
    return str(record["game"]), str(record["date"])


def draw_value(record: dict[str, object]) -> tuple[int, ...]:
    return tuple(int(v) for v in record["numbers"])


def validate_draw(record: dict[str, object]) -> None:
    game = str(record["game"])
    if game not in GAMES:
        raise ValueError(f"unknown game: {game}")
    numbers = draw_value(record)
    if len(numbers) != 6 or len(set(numbers)) != 6:
        raise ValueError(f"{draw_key(record)}: expected 6 distinct numbers")
    if not all(1 <= value <= GAMES[game] for value in numbers):
        raise ValueError(f"{draw_key(record)}: number outside game pool")


def validate_manifest(manifest: dict[str, object]) -> list[dict[str, object]]:
    if manifest.get("schema_version") != 1:
        raise ValueError("manifest schema_version must be 1")
    sources = {s["id"] for s in manifest["sources"]}
    records = manifest["draws"]
    if len(records) != 28:
        raise ValueError(f"manifest: expected 28 draws, got {len(records)}")
    seen: set[tuple[str, str]] = set()
    for record in records:
        validate_draw(record)
        key = draw_key(record)
        if key in seen:
            raise ValueError(f"manifest duplicate: {key}")
        seen.add(key)
        if set(record["source_ids"]) != sources:
            raise ValueError(f"{key}: expected both registered source IDs")
        if record["status"] != "two_source_verified":
            raise ValueError(f"{key}: unexpected audit status")
    return records


def validate_inputs(
    manifest: dict[str, object], records: list[dict[str, object]]
) -> tuple[list[dict[str, object]], dict[tuple[str, str], dict[str, str]]]:
    cutoff = str(manifest["confirmation_cutoff"])
    previous_latest = str(manifest["previous_latest_draw_date"])
    latest = str(manifest["latest_draw_date"])
    expected_counts = manifest["expected_row_counts"]
    expected_batch = {draw_key(r): draw_value(r) for r in records}
    confirmation_sets = []

    for filename, date_column in DRAW_FILES.items():
        path = DATASET_DIR / filename
        require_crlf(path)
        rows = dated_rows(path, date_column)
        if len(rows) != expected_counts[filename]:
            raise ValueError(
                f"{filename}: expected {expected_counts[filename]} rows, "
                f"got {len(rows)}"
            )
        parsed = [draw_record(row, date_column) for row in rows]
        for record in parsed:
            validate_draw(record)
        keys = [draw_key(record) for record in parsed]
        if len(keys) != len(set(keys)):
            raise ValueError(f"{filename}: duplicate game/date key")
        if max(record["date"] for record in parsed) != latest:
            raise ValueError(f"{filename}: unexpected latest draw date")
        batch = {
            draw_key(record): draw_value(record)
            for record in parsed
            if record["date"] > previous_latest
        }
        if batch != expected_batch:
            raise ValueError(f"{filename}: batch differs from manifest")
        prefix_hash = exploration_prefix_sha256(path, date_column, cutoff)
        if prefix_hash != EXPLORATION_PREFIX_SHA256[filename]:
            raise ValueError(f"{filename}: frozen exploration prefix changed")
        confirmation_sets.append(
            {
                draw_key(record): draw_value(record)
                for record in parsed
                if record["date"] > cutoff
            }
        )

        if filename.endswith("_audited.csv"):
            by_key = {draw_key(draw_record(r, date_column)): r for r in rows}
            for record in records:
                row = by_key[draw_key(record)]
                if [row["Source1"], row["Source2"]] != record["source_ids"]:
                    raise ValueError(f"{draw_key(record)}: audit sources differ")
                if row["Status"] != record["status"]:
                    raise ValueError(f"{draw_key(record)}: audit status differs")

    if not all(item == confirmation_sets[0] for item in confirmation_sets[1:]):
        raise ValueError("draw CSVs disagree on confirmation rows")

    canonical_rows = dated_rows(DATASET_DIR / "data_draws_1yr.csv", "Date")
    confirmation = [
        draw_record(row, "Date") for row in canonical_rows if row["Date"] > cutoff
    ]

    astro_path = DATASET_DIR / ASTRO_FILE
    require_crlf(astro_path)
    if (
        exploration_prefix_sha256(astro_path, ASTRO_DATE, cutoff)
        != EXPLORATION_PREFIX_SHA256[ASTRO_FILE]
    ):
        raise ValueError(f"{ASTRO_FILE}: frozen exploration prefix changed")
    astro_rows = dated_rows(astro_path, ASTRO_DATE)
    if len(astro_rows) != expected_counts[ASTRO_FILE]:
        raise ValueError(f"{ASTRO_FILE}: unexpected dated row count")
    astro_by_key = {(row["Game"], row[ASTRO_DATE]): row for row in astro_rows}
    if len(astro_by_key) != len(astro_rows):
        raise ValueError(f"{ASTRO_FILE}: duplicate game/date key")
    for record in confirmation:
        key = draw_key(record)
        if key not in astro_by_key:
            raise ValueError(f"{ASTRO_FILE}: missing {key}")
        row = astro_by_key[key]
        expected_mean = sum(draw_value(record)) / (6 * GAMES[str(record["game"])])
        if f"{expected_mean:.4f}" != row[MEAN_DRAWN]:
            raise ValueError(f"{ASTRO_FILE}: mean-drawn mismatch for {key}")
        if row[KP_DRAW] or row[KP_DAILY]:
            raise ValueError(f"{ASTRO_FILE}: expected blank Kp for {key}")
    return confirmation, astro_by_key


def workbook_invariants() -> dict[str, object]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for workbook verification") from exc

    formulas_book = load_workbook(WORKBOOK, data_only=False, read_only=False)
    expected_sheets = [
        "Read Me",
        "Draws",
        "Frequency",
        "Hot-Cold",
        "Moon-Sun Test",
        "EV Calculator",
        "Protocol",
        "Future Draws",
    ]
    if formulas_book.sheetnames != expected_sheets:
        raise ValueError("workbook sheet inventory changed")
    formula_count = sum(
        1
        for sheet in formulas_book.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
        or (isinstance(cell.value, str) and cell.value.startswith("="))
    )
    if formula_count != 786:
        raise ValueError(f"workbook: expected 786 formulas, got {formula_count}")

    values_book = load_workbook(WORKBOOK, data_only=True, read_only=False)
    errors = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in values_book.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "e"
        or (isinstance(cell.value, str) and cell.value.startswith("#"))
    ]
    if errors:
        raise ValueError(f"workbook cached-value errors: {errors[:5]}")
    if formulas_book["Draws"].max_row != 253:
        raise ValueError("workbook Draws sheet must contain 252 data rows")
    return {
        "sha256": sha256(WORKBOOK),
        "sheets": len(expected_sheets),
        "formulas": formula_count,
        "cached_value_errors": 0,
        "draw_rows": formulas_book["Draws"].max_row - 1,
    }


def pearson(x: list[float], y: list[float]) -> float:
    if len(x) != len(y) or len(x) < 3:
        raise ValueError("pearson inputs must have equal length >= 3")
    x_mean = sum(x) / len(x)
    y_mean = sum(y) / len(y)
    x_centered = [value - x_mean for value in x]
    y_centered = [value - y_mean for value in y]
    denominator = math.sqrt(
        sum(value * value for value in x_centered)
        * sum(value * value for value in y_centered)
    )
    if denominator == 0:
        raise ValueError("pearson input has zero variance")
    return sum(a * b for a, b in zip(x_centered, y_centered)) / denominator


def within_game_predictor(
    confirmation: list[dict[str, object]],
    astro: dict[tuple[str, str], dict[str, str]],
    covariate: str,
) -> tuple[list[float], list[float]]:
    predictor: list[float] = []
    response: list[float] = []
    for game in GAMES:
        rows = [record for record in confirmation if record["game"] == game]
        means = [float(astro[draw_key(record)][MEAN_DRAWN]) for record in rows]
        mean = sum(means) / len(means)
        std = math.sqrt(sum((value - mean) ** 2 for value in means) / len(means))
        predictor.extend((value - mean) / std for value in means)
        response.extend(float(astro[draw_key(record)][covariate]) for record in rows)
    return predictor, response


def chi_square(record_list: list[dict[str, object]], pool: int) -> float:
    expected = len(record_list) * 6 / pool
    counts = Counter(value for record in record_list for value in draw_value(record))
    return sum(
        (counts[value] - expected) ** 2 / expected
        for value in range(1, pool + 1)
    )


def binomial_two_sided(k: int, n: int, probability: float) -> float:
    probabilities = [
        math.comb(n, value)
        * probability**value
        * (1 - probability) ** (n - value)
        for value in range(n + 1)
    ]
    observed = probabilities[k]
    return min(1.0, sum(value for value in probabilities if value <= observed + 1e-15))


def run_monitoring(
    manifest: dict[str, object],
    confirmation: list[dict[str, object]],
    astro: dict[tuple[str, str], dict[str, str]],
    manifest_path: Path,
) -> dict[str, object]:
    seed = int(manifest["seed"])
    mc_trials = int(manifest["mc_trials"])
    permutation_trials = int(manifest["permutation_trials"])
    family_m = int(manifest["registered_family_m"])
    threshold = 0.05 / family_m
    rng = random.Random(seed)
    flags: list[str] = []
    chi_results: dict[str, object] = {}

    for game, pool in GAMES.items():
        rows = [record for record in confirmation if record["game"] == game]
        observed = chi_square(rows, pool)
        exceedances = 0
        for _ in range(mc_trials):
            simulated = [
                {
                    "date": "",
                    "game": game,
                    "numbers": rng.sample(range(1, pool + 1), 6),
                }
                for _ in rows
            ]
            exceedances += chi_square(simulated, pool) >= observed - 1e-12
        p_value = (exceedances + 1) / (mc_trials + 1)
        flagged = p_value < threshold
        test_id = f"chi_square:{game}"
        if flagged:
            flags.append(test_id)
        chi_results[game] = {
            "n": len(rows),
            "chi2": round(observed, 6),
            "p_mc": round(p_value, 6),
            "mc_exceedances": exceedances,
            "flag": flagged,
        }

    permutation_results: dict[str, object] = {}
    for test_id, covariate in (
        ("mean_drawn_vs_moon_altitude", MOON_ALT),
        ("mean_drawn_vs_moon_illumination", MOON_ILLUM),
    ):
        predictor, response = within_game_predictor(
            confirmation, astro, covariate
        )
        observed = pearson(predictor, response)
        exceedances = 0
        for _ in range(permutation_trials):
            permuted = response.copy()
            rng.shuffle(permuted)
            exceedances += abs(pearson(predictor, permuted)) >= abs(observed) - 1e-15
        p_value = (exceedances + 1) / (permutation_trials + 1)
        flagged = p_value < threshold
        if flagged:
            flags.append(f"pearson:{test_id}")
        permutation_results[test_id] = {
            "r": round(observed, 6),
            "p": round(p_value, 6),
            "permutation_exceedances": exceedances,
            "n": len(predictor),
            "flag": flagged,
        }

    permutation_results["mean_drawn_vs_kp"] = {
        "status": "not_computable",
        "reason": "Kp blank for confirmation draws",
    }

    grand_rows = [
        record for record in confirmation if record["game"] == "Grand Lotto 6/55"
    ]
    hits = sum(45 in draw_value(record) for record in grand_rows)
    binomial_p = binomial_two_sided(hits, len(grand_rows), 6 / 55)
    binomial_flag = binomial_p < threshold
    if binomial_flag:
        flags.append("binomial:Grand Lotto 6/55:number45")

    inputs = [
        DATASET_DIR / filename
        for filename in (*DRAW_FILES.keys(), ASTRO_FILE)
    ]
    input_hashes = {
        str(path.relative_to(ROOT)): sha256(path) for path in inputs
    }
    input_hashes[str(WORKBOOK.relative_to(ROOT))] = sha256(WORKBOOK)
    input_hashes[str(manifest_path.relative_to(ROOT))] = sha256(manifest_path)

    sizes = {
        game: sum(record["game"] == game for record in confirmation)
        for game in GAMES
    }
    sizes["total"] = len(confirmation)

    return {
        "_meta": {
            "schema_version": 1,
            "script": "src/pcso_weekly_update.py",
            "registration": "datasets/pcso-lotto/DATASET.md section 6",
            "run_date": manifest["run_date"],
            "pipeline": "pcso-weekly-update",
            "grade": "G0 exploratory forward monitoring",
            "confirmation_cutoff": manifest["confirmation_cutoff"],
            "seed_scheme": (
                f"single random.Random({seed}) stream in published test order"
            ),
            "null_simulator": (
                "random.sample(range(1, P+1), 6) independently per draw"
            ),
            "add_one_correction": True,
            "test_order": [
                *[f"chi_square:{game}" for game in GAMES],
                "pearson:mean_drawn_vs_moon_altitude",
                "pearson:mean_drawn_vs_moon_illumination",
                "pearson:mean_drawn_vs_kp",
                "binomial:Grand Lotto 6/55:number45",
            ],
            "input_sha256": input_hashes,
            "exploration_prefix_sha256": EXPLORATION_PREFIX_SHA256,
            "workbook": workbook_invariants(),
            "audit_correction": {
                "supersedes_unreproducible_artifact_sha256": (
                    "a9e1424b3f2f8d752867f760b38b69f5b9ad39de2f787b6741beb26f6dcc7d50"
                ),
                "reason": (
                    "the original untracked JSON had no generating script; "
                    "this rerun applies the dataset card's declared null and "
                    "publishes the RNG stream contract"
                ),
                "verdict_changed": False,
            },
        },
        "confirmation_sizes": sizes,
        "registered_family_m": family_m,
        "computable_tests": 8,
        "bonferroni_threshold": round(threshold, 6),
        "seed": seed,
        "tests": {
            f"chi_square_mc_{mc_trials}": chi_results,
            f"pearson_permutation_{permutation_trials}": permutation_results,
            "binomial_655_number45": {
                "n_655": len(grand_rows),
                "k": hits,
                "expected": round(len(grand_rows) * 6 / 55, 6),
                "p_two_sided": round(binomial_p, 6),
                "flag": binomial_flag,
            },
        },
        "flags": flags,
        "note": (
            "The m=9 threshold is within-look only. Weekly cumulative looks are "
            "exploratory monitoring without a registered sequential alpha-spending "
            "rule; any flag requires a separately registered fresh replication. "
            "Kp remains non-computable pending GFZ backfill."
        ),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manifest_path = args.manifest.resolve()
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    records = validate_manifest(manifest)
    confirmation, astro = validate_inputs(manifest, records)
    result = run_monitoring(manifest, confirmation, astro, manifest_path)
    output = args.out.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(result, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )
    print(
        f"validated {len(records)} new draws; "
        f"confirmation n={len(confirmation)}; flags={len(result['flags'])}; "
        f"wrote {output}"
    )


if __name__ == "__main__":
    main()
